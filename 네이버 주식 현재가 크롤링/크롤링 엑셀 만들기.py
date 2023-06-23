import openpyxl

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('data')

# 3) 데이터 추가하기
ws['A1'] = '종목'
ws['B1'] = '현재가'

ws['A2'] = 'KT'
ws['A3'] = 'LG전자'
ws['A4'] = '삼성전자'


# 4) 엑셀 저장하기
wb.save(r'C:\gb_0900_kdh\python\네이버 주식 현재가 크롤링\주식가_data.xlsx')